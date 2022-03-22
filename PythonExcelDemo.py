import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Gulshan\\Documents\\Book1.xlsx")
sheet = book.active
Dict={}
# print(sheet.max_column)
row = sheet.max_row
for j in range(2,sheet.max_row+1):
    print(j)
    for i in range(1,sheet.max_column+1):
        Dict = [sheet.cell(row=1 , column=i).value] = sheet.cell(row=j, column=i).value
print(Dict)